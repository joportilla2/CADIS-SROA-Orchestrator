from __future__ import annotations

"""
Author: Omar Portilla Jaimes
Email: jorge.portilla2@unipamplona.edu.co
"""

from typing import Any, Dict, Iterable, List, Sequence, Tuple, Optional
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def to_excel_value(v: Any) -> Any:
    """Convert Python objects to Excel-safe values (openpyxl-compatible)."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, Path):
        return str(v)
    if isinstance(v, (dict, list, tuple, set)):
        return json.dumps(v, ensure_ascii=False, sort_keys=True)
    if hasattr(v, "item") and callable(getattr(v, "item")):
        try:
            return v.item()
        except Exception:
            pass
    return str(v)

def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for j, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v)
            widths[j] = min(max_width, max(widths.get(j, 0), len(s) + 2))
    for j, w in widths.items():
        ws.column_dimensions[get_column_letter(j)].width = w

def write_kv_sheet(ws: Worksheet, items: Sequence[Tuple[str, Any]]) -> None:
    ws.append(["key", "value"])
    for k, v in items:
        ws.append([to_excel_value(k), to_excel_value(v)])
    _autosize(ws)

def write_table_sheet(ws: Worksheet, columns: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    ws.append([to_excel_value(c) for c in columns])
    for r in rows:
        ws.append([to_excel_value(x) for x in r])
    _autosize(ws)

def write_dict_rows_sheet(ws: Worksheet, rows: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> None:
    if not rows:
        # Write an empty sheet with headers if provided
        cols = columns or []
        ws.append([to_excel_value(c) for c in cols])
        _autosize(ws)
        return

    cols = columns
    if cols is None:
        # Stable order for reproducibility
        keyset = set()
        for r in rows:
            keyset.update(r.keys())
        cols = sorted(keyset)

    ws.append([to_excel_value(c) for c in cols])
    for row in rows:
        ws.append([to_excel_value(row.get(c, None)) for c in cols])
    _autosize(ws)

def save_workbook(path: str, sheets: Dict[str, Dict[str, Any]]) -> str:
    """Create an .xlsx file with named sheets.

    sheets[sheet_name] must include one of:
      - {'type': 'kv', 'items': [(k,v), ...]}
      - {'type': 'table', 'columns': [...], 'rows': [[...], ...]}
      - {'type': 'dict_rows', 'rows': [dict,...], 'columns': [...] (optional)}
    """
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    for name, spec in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        t = spec.get("type")
        if t == "kv":
            write_kv_sheet(ws, spec.get("items", []))
        elif t == "table":
            write_table_sheet(ws, spec.get("columns", []), spec.get("rows", []))
        elif t == "dict_rows":
            write_dict_rows_sheet(ws, spec.get("rows", []), spec.get("columns"))
        else:
            raise ValueError(f"Unknown sheet type: {t}")
    wb.save(path)
    return path
